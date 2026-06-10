import pandas as pd
import uuid
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

file_path = "Laporan Harian Email GBI 01.04.26.xlsx"
xls = pd.ExcelFile(file_path)

# --- 1. SETUP MASTER DATA (SYSTEM DEFAULTS) ---
org_id = str(uuid.uuid4())
loc_gbi_id = str(uuid.uuid4())
tx_date = datetime.date(2026, 4, 1)

locations_data = [{
    "location_id": loc_gbi_id, "customer_id": None, "location_type": 1,
    "location_name": "GF GBI", "location_city": "Bandung", "location_address": "GBI", "location_province": "Jawa Barat",
    "organization_id": org_id
}]

coa_data = [
    {"coa_id": str(uuid.uuid4()), "coa_code": "1001", "coa_name": "Kas Apotek", "coa_type": "Asset", "normal_balance": "Debit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "1002", "coa_name": "Kas di Bank (EDC/QRIS)", "coa_type": "Asset", "normal_balance": "Debit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "1003", "coa_name": "Persediaan Obat", "coa_type": "Asset", "normal_balance": "Debit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "2001", "coa_name": "Hutang Usaha PBF", "coa_type": "Liability", "normal_balance": "Credit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "4001", "coa_name": "Pendapatan Penjualan", "coa_type": "Revenue", "normal_balance": "Credit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "5001", "coa_name": "Harga Pokok Penjualan", "coa_type": "Expense", "normal_balance": "Debit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "6001", "coa_name": "Biaya Operasional", "coa_type": "Expense", "normal_balance": "Debit"},
    {"coa_id": str(uuid.uuid4()), "coa_code": "6002", "coa_name": "Selisih Kas (Kurang/Lebih)", "coa_type": "Expense", "normal_balance": "Debit"},
]
coa_map = {c["coa_code"]: c for c in coa_data}

b_mandiri_id = str(uuid.uuid4())
b_bca_id = str(uuid.uuid4())
b_muamalat_id = str(uuid.uuid4())
bank_data = [
    {"id": b_mandiri_id, "bank_name": "Bank Mandiri", "bank_account_number": "13000-MANDIRI", "bank_account_holder": "Apotek GBI", "organization_id": org_id},
    {"id": b_bca_id, "bank_name": "Bank BCA", "bank_account_number": "01400-BCA", "bank_account_holder": "Apotek GBI", "organization_id": org_id},
    {"id": b_muamalat_id, "bank_name": "Bank Muamalat", "bank_account_number": "14700-MUAMALAT", "bank_account_holder": "Apotek GBI", "organization_id": org_id},
]

fiscal_period_id = str(uuid.uuid4())
fiscal_period_data = [{
    "fiscal_period_id": fiscal_period_id, "period_name": "April 2026", "period_month": 4, "period_year": 2026,
    "start_date": datetime.date(2026, 4, 1), "end_date": datetime.date(2026, 4, 30), "status": "OPEN", "organization_id": org_id
}]

# --- 2. EXTRACT MASTER DATA DARI EXCEL ---
customers_dict = {}
pic_dict = {}
if 'PARETO KARYAWAN' in xls.sheet_names:
    df_pareto = pd.read_excel(file_path, sheet_name='PARETO KARYAWAN')
    for idx, row in df_pareto.iterrows():
        if idx < 8: continue
        cust_name = str(row.iloc[5]).strip()
        if cust_name and cust_name != 'nan' and cust_name not in customers_dict:
            customers_dict[cust_name] = str(uuid.uuid4())
            
        emp_name = str(row.iloc[2]).strip()
        if emp_name and emp_name != 'nan' and emp_name not in pic_dict:
            pic_dict[emp_name] = str(uuid.uuid4())

if "ANONIM" not in customers_dict: customers_dict["ANONIM"] = str(uuid.uuid4())

customers_data = [{"customer_id": cid, "customer_name": name, "customer_code": f"CUST-{str(cid)[:8]}", "customer_email": None} for name, cid in customers_dict.items()]
pic_data = [{"pic_id": pid, "pic_name": name, "pic_phone_number": None, "pic_email": None, "customer_id": None, "supplier_id": None} for name, pid in pic_dict.items()]

suppliers_dict = {}
valid_purchase_keywords = ["PEMBELIAN PBF", "PEMBELIAN OUTLET", "MUTASI MASUK"]
purchase_sheets = [s for s in xls.sheet_names if any(keyword in s for keyword in valid_purchase_keywords)]
for sheet in purchase_sheets:
    df_p = pd.read_excel(file_path, sheet_name=sheet)
    for idx, row in df_p.iterrows():
        row_str = " ".join([str(x) for x in row.values if pd.notna(x)])
        if "No Faktur :" in row_str:
            match_sup = re.search(r'\((.*?)\)', row_str)
            supplier = match_sup.group(1).strip() if match_sup else "UNKNOWN"
            if supplier not in suppliers_dict:
                suppliers_dict[supplier] = str(uuid.uuid4())

suppliers_data = [{"supplier_id": sid, "supplier_name": name, "supplier_type": "PBF", "supplier_code": f"SUP-{str(sid)[:8]}", "supplier_email": None, "supplier_city": None, "supplier_address": None} for name, sid in suppliers_dict.items()]


# --- 3. DYNAMIC TRANSACTION & JOURNAL EXTRACTION ---
products_dict = {}
po_headers, po_details, stock_movements, journal_entries, journal_lines = [], [], [], [], []
stock_balances = {}

def create_jline(jid, coa_code, entry_type, amount, pay_method=None, b_id=None, desc=""):
    coa = coa_map[coa_code]
    return {
        "line_id": str(uuid.uuid4()), "journal_id": jid, 
        "coa_id": coa["coa_id"], "coa_code": coa["coa_code"], "coa_name": coa["coa_name"], "coa_type": coa["coa_type"],
        "entry_type": entry_type, "debit_amount": amount if entry_type == "DR" else 0, "credit_amount": amount if entry_type == "CR" else 0,
        "payment_method": pay_method, "bank_id": b_id, "description": desc, "organization_id": org_id
    }

# A. PROSES PENJUALAN (SALES)
for sheet in ['PENJUALAN SHIFT 1', 'PENJUALAN SHIFT 2']:
    if sheet not in xls.sheet_names: continue
    df_s = pd.read_excel(file_path, sheet_name=sheet)
    curr_po_num, curr_po_line_id = None, None
    curr_cust_id = customers_dict["ANONIM"]
    curr_pay_method = "Tunai/Cash"
    
    for idx, row in df_s.iterrows():
        row_str = " ".join([str(x) for x in row.values if pd.notna(x)])
        match_inv = re.search(r'(SL-\d{6}-\d{5})', row_str)
        match_pay = re.search(r'\((.*?)\)', row_str)
        if match_inv:
            curr_po_num = match_inv.group(1)
            curr_po_line_id = str(uuid.uuid4())
            curr_cust_id = customers_dict["ANONIM"]
            match_pay = re.search(re.escape(curr_po_num) + r'.*?\((.*?)\)', row_str)
            print('match_pay',match_pay)
            if match_pay:
                curr_pay_method = match_pay.group(1).strip() # Hasil: Bebas/EDC-Qris Muamalat
                print('curr_pay_method',curr_pay_method)
            else:
                curr_pay_method = "Tunai/Cash" # Default jika kebetulan kasir lupa input kurung
            
            po_headers.append({
                "purchase_order_line_number": curr_po_line_id, "purchase_order_number": curr_po_num,
                "purchase_order_type": 1, "supplier_id": None, "customer_id": curr_cust_id, "pic_id": None,
                "tax_id": None, "location_id": loc_gbi_id, "purchase_order_status": 2, "purchase_order_total": 0,
                "payment_method": curr_pay_method # Temporary field for journal engine
            })
            
        elif "Nama :" in row_str and curr_po_num:
            try:
                name_part = row_str.split("Nama :")[1].split("Umur :")[0].strip()
                if name_part and name_part in customers_dict:
                    po_headers[-1]["customer_id"] = customers_dict[name_part]
                # Extrak Pembayaran (Bebas/Tunai, EDC Mandiri, dll)
                pay_match = re.search(r'\((.*?)\)', row_str)
                if pay_match:
                    po_headers[-1]["payment_method"] = pay_match.group(1).strip()
            except: pass
            
        elif curr_po_num and pd.notna(row.iloc[0]) and isinstance(row.iloc[0], str) and row.iloc[0] not in ["No", "Nama Barang", "Tanggal :", "Outlet :", "Shift :", "LAPORAN PENJUALAN", "Diskon Faktur :"]:
            try:
                if isinstance(row.iloc[1], (int, float)):
                    prod_name = str(row.iloc[0]).strip()
                    if prod_name not in products_dict: products_dict[prod_name] = str(uuid.uuid4())
                    pid = products_dict[prod_name]
                    
                    qty = float(row.iloc[1])
                    hpp = float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
                    price = float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
                    subtotal = qty * price
                    
                    po_headers[-1]["purchase_order_total"] += subtotal
                    
                    po_details.append({
                        "purchase_order_line_number": str(uuid.uuid4()), "purchase_order_number": curr_po_num,
                        "product_id": pid, "purchase_order_quantity": qty, "purchase_order_hpp": hpp,
                        "purchase_order_selling_price": price, "purchase_order_quantity_unit": 1, "tax_id": None
                    })
                    
                    if pid not in stock_balances: stock_balances[pid] = 0
                    stock_movements.append({
                        "stock_movement_line_number": str(uuid.uuid4()), "stock_movement_stock_id": pid,
                        "stock_movement_location_id": loc_gbi_id, "stock_movement_purchase_order_number": curr_po_num,
                        "stock_movement_type": 2, "stock_movement_quantity": qty
                    })
                    stock_balances[pid] -= qty
            except: pass

# B. PROSES PEMBELIAN (PURCHASE PBF)
for sheet in purchase_sheets:
    df_p = pd.read_excel(file_path, sheet_name=sheet)
    curr_po_num, curr_po_line_id = None, None
    
    for idx, row in df_p.iterrows():
        row_str = " ".join([str(x) for x in row.values if pd.notna(x)])
        if "No Faktur :" in row_str:
            match_inv = re.search(r'No Faktur :\s*([\w/-]+)', row_str)
            curr_po_num = match_inv.group(1) if match_inv else "UNKNOWN-PO"
            curr_po_line_id = str(uuid.uuid4())
            match_sup = re.search(r'\((.*?)\)', row_str)
            sup_name = match_sup.group(1).strip() if match_sup else "UNKNOWN"
            
            po_headers.append({
                "purchase_order_line_number": curr_po_line_id, "purchase_order_number": curr_po_num,
                "purchase_order_type": 2, "supplier_id": suppliers_dict.get(sup_name, None),
                "customer_id": None, "pic_id": None, "tax_id": None, "location_id": loc_gbi_id,
                "purchase_order_status": 1, "purchase_order_total": 0, "payment_method": "Hutang PBF"
            })
            
        elif curr_po_num and pd.notna(row.iloc[4]) and isinstance(row.iloc[4], str) and "Sub Total:" not in row_str and "PPN:" not in row_str:
            try:
                prod_name = str(row.iloc[4]).strip()
                if prod_name and prod_name not in ["Nama Barang", "nan", "Diskon Faktur:"]:
                    if prod_name not in products_dict: products_dict[prod_name] = str(uuid.uuid4())
                    pid = products_dict[prod_name]
                    
                    qty = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                    harga_beli = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
                    subtot = qty * harga_beli
                    po_headers[-1]["purchase_order_total"] += subtot
                    
                    po_details.append({
                        "purchase_order_line_number": str(uuid.uuid4()), "purchase_order_number": curr_po_num,
                        "product_id": pid, "purchase_order_quantity": qty, "purchase_order_hpp": harga_beli,
                        "purchase_order_selling_price": 0, "purchase_order_quantity_unit": 1, "tax_id": None
                    })
                    
                    if pid not in stock_balances: stock_balances[pid] = 0
                    stock_movements.append({
                        "stock_movement_line_number": str(uuid.uuid4()), "stock_movement_stock_id": pid,
                        "stock_movement_location_id": loc_gbi_id, "stock_movement_purchase_order_number": curr_po_num,
                        "stock_movement_type": 1, "stock_movement_quantity": qty
                    })
                    stock_balances[pid] += qty
            except: pass


# C. BUILD JOURNALS DARI DATA PO YANG SUDAH DIEKSTRAK
for po in po_headers:
    is_sales = po["purchase_order_type"] == 1
    tx_type = "SALES" if is_sales else "PURCHASE"
    src_type = "INVOICE" if is_sales else "FAKTUR_PBF"
    total_amount = po["purchase_order_total"]
    
    # Hitung total HPP dari details
    rel_details = [d for d in po_details if d["purchase_order_number"] == po["purchase_order_number"]]
    total_hpp = sum(d["purchase_order_hpp"] * d["purchase_order_quantity"] for d in rel_details)
    pay_method = po.get("payment_method", "")
    
    jid = str(uuid.uuid4())
    journal_entries.append({
        "journal_id": jid, "journal_date": tx_date, "journal_type": tx_type,
        "source_document_type": src_type, "source_document_id": po["purchase_order_number"],
        "stock_movement_id": None, "description": f"{tx_type} Dokumen {po['purchase_order_number']}",
        "fiscal_period_id": fiscal_period_id, "location_id": loc_gbi_id, "status": "POSTED", "organization_id": org_id
    })
    
    if is_sales:
        # DETEKSI METODE PEMBAYARAN CERDAS
        pay_str = pay_method.upper()
        if any(k in pay_str for k in ["EDC", "QRIS", "BANK", "TRANSFER", "DEBIT"]):
            b_id = b_mandiri_id # Default fallback
            if "BCA" in pay_str: b_id = b_bca_id
            elif "MUAMALAT" in pay_str: b_id = b_muamalat_id
            journal_lines.append(create_jline(jid, "1002", "DR", total_amount, pay_method, b_id, "Penerimaan EDC/QRIS"))
        else:
            journal_lines.append(create_jline(jid, "1001", "DR", total_amount, pay_method, None, "Penerimaan Tunai Laci"))
            
        journal_lines.append(create_jline(jid, "4001", "CR", total_amount, desc="Pendapatan Penjualan"))
        journal_lines.append(create_jline(jid, "5001", "DR", total_hpp, desc="HPP"))
        journal_lines.append(create_jline(jid, "1003", "CR", total_hpp, desc="Pengurangan Stok Obat"))
    else:
        journal_lines.append(create_jline(jid, "1003", "DR", total_amount, desc="Penambahan Stok"))
        journal_lines.append(create_jline(jid, "2001", "CR", total_amount, desc="Hutang PBF"))
        
    # Cleanup temporary dict keys
    po.pop("payment_method", None)


# D. PROSES BIAYA OPERASIONAL (EXPENSES)
if 'BIAYA' in xls.sheet_names:
    df_b = pd.read_excel(file_path, sheet_name='BIAYA')
    for idx, row in df_b.iterrows():
        # Asumsikan kolom 2 adalah Nomor Bukti, kolom 3 keterangan, kolom 4 jumlah
        if pd.notna(row.iloc[2]) and "KO-" in str(row.iloc[2]):
            try:
                doc_num = str(row.iloc[2]).strip()
                desc = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else "Biaya Operasional"
                amount = float(row.iloc[4])
                
                jid = str(uuid.uuid4())
                journal_entries.append({
                    "journal_id": jid, "journal_date": tx_date, "journal_type": "EXPENSE",
                    "source_document_type": "KAS_KELUAR", "source_document_id": doc_num,
                    "stock_movement_id": None, "description": desc,
                    "fiscal_period_id": fiscal_period_id, "location_id": loc_gbi_id, "status": "POSTED", "organization_id": org_id
                })
                journal_lines.append(create_jline(jid, "6001", "DR", amount, desc=desc))
                journal_lines.append(create_jline(jid, "1001", "CR", amount, "Tunai", None, "Kas Keluar"))
            except: pass

# E. PROSES SELISIH TUTUP SHIFT
if 'TUTUP SHIFT' in xls.sheet_names:
    df_t = pd.read_excel(file_path, sheet_name='TUTUP SHIFT')
    for idx, row in df_t.iterrows():
        row_str = " ".join([str(x) for x in row.values if pd.notna(x)])
        # Ekstrak data jika ada teks kurang bayar
        if "Kurang Bayar" in row_str or "Selisih Kurang" in row_str:
            try:
                # Cari angka di kolom terdekat
                amount = float(re.sub(r'[^\d.]', '', str(row.iloc[-1])))
                if amount > 0:
                    jid = str(uuid.uuid4())
                    journal_entries.append({
                        "journal_id": jid, "journal_date": tx_date, "journal_type": "RECONCILIATION",
                        "source_document_type": "TUTUP_SHIFT", "source_document_id": f"RECON-{str(uuid.uuid4())[:8]}",
                        "stock_movement_id": None, "description": "Selisih Kurang Bayar Shift",
                        "fiscal_period_id": fiscal_period_id, "location_id": loc_gbi_id, "status": "POSTED", "organization_id": org_id
                    })
                    journal_lines.append(create_jline(jid, "6002", "DR", amount, desc="Beban Selisih Kas"))
                    journal_lines.append(create_jline(jid, "1001", "CR", amount, desc="Penyesuaian Fisik Kas Laci"))
            except: pass

# --- 4. FORMATTING & SAVE TO EXCEL ---
products_data = [{"id": pid, "product_name": name, "product_code": f"PRD-{str(pid)[:8]}"} for name, pid in products_dict.items()]
stock_data = [{"id": str(uuid.uuid4()), "stock_product_id": pid, "stock_location_id": loc_gbi_id, "stock_quantity_available": qty} for pid, qty in stock_balances.items()]

output_path = "ERP_Final_Dynamic_Export.xlsx" # Simpan di direktori lokal Anda
wb = openpyxl.Workbook()
wb.remove(wb.active)

datasets = {
    "locations": pd.DataFrame(locations_data),
    "chart_of_accounts": pd.DataFrame(coa_data),
    "fiscal_periods": pd.DataFrame(fiscal_period_data),
    "bank": pd.DataFrame(bank_data),
    "pic": pd.DataFrame(pic_data),
    "customers": pd.DataFrame(customers_data),
    "suppliers": pd.DataFrame(suppliers_data),
    "products": pd.DataFrame(products_data),
    "purchase_order_header": pd.DataFrame(po_headers),
    "purchase_order_detail": pd.DataFrame(po_details),
    "stock": pd.DataFrame(stock_data),
    "stock_movement": pd.DataFrame(stock_movements),
    "journal_entries": pd.DataFrame(journal_entries),
    "journal_entry_lines": pd.DataFrame(journal_lines)
}

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for sheet_name, df in datasets.items():
    ws = wb.create_sheet(title=sheet_name)
    if df.empty: continue
    
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style
        
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

wb.save(output_path)
print(f"File berhasi diproses dan disimpan ke: {output_path}")