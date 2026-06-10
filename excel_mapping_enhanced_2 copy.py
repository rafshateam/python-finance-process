import pandas as pd
import uuid
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# --- 1. INISIALISASI & SETUP MASTER DATA BERSKALA ERD ---
file_path = "Combined_Output.xlsx"
try:
    xls = pd.ExcelFile(file_path)
    print(f"File ditemukan! Membaca {len(xls.sheet_names)} sheets...")
except Exception as e:
    print(f"Gagal membaca file {file_path}. Error: {e}")
    exit()

org_id = "ORG-GBI-001"
loc_gbi_id = str(uuid.uuid4())
now = datetime.datetime.now()
import_batch_id = str(uuid.uuid4())

# Tabel: LOCATIONS
locations_data = [{
    "id": loc_gbi_id, "location_id": 1, "location_name": "GF GBI", 
    "location_address": "GBI Margasari", "location_city": "Bandung"
}]

# Tabel: CHART_OF_ACCOUNTS
coa_data = [
    {"id": str(uuid.uuid4()), "coa_code": "1001", "coa_name": "Kas Apotek", "coa_type": "Asset", "normal_balance": "Debit"},
    {"id": str(uuid.uuid4()), "coa_code": "1002", "coa_name": "Kas di Bank (EDC/QRIS/TF)", "coa_type": "Asset", "normal_balance": "Debit"},
    {"id": str(uuid.uuid4()), "coa_code": "1003", "coa_name": "Persediaan Obat", "coa_type": "Asset", "normal_balance": "Debit"},
    {"id": str(uuid.uuid4()), "coa_code": "4001", "coa_name": "Pendapatan Penjualan", "coa_type": "Revenue", "normal_balance": "Credit"},
    {"id": str(uuid.uuid4()), "coa_code": "5001", "coa_name": "Harga Pokok Penjualan", "coa_type": "Expense", "normal_balance": "Debit"},
    {"id": str(uuid.uuid4()), "coa_code": "6001", "coa_name": "Biaya Operasional", "coa_type": "Expense", "normal_balance": "Debit"}
]
coa_map = {c["coa_code"]: c for c in coa_data}

# Tabel: BANK & PAYMENT_METHODS
b_mandiri_id, b_bca_id, b_muamalat_id = str(uuid.uuid4()), str(uuid.uuid4()), str(uuid.uuid4())
bank_data = [
    {"id": b_mandiri_id, "bank_id": 1, "bank_name": "Bank Mandiri", "bank_account_number": "13000", "bank_account_holder": "Apotek GBI"},
    {"id": b_bca_id, "bank_id": 2, "bank_name": "Bank BCA", "bank_account_number": "01400", "bank_account_holder": "Apotek GBI"},
    {"id": b_muamalat_id, "bank_id": 3, "bank_name": "Bank Muamalat", "bank_account_number": "14700", "bank_account_holder": "Apotek GBI"}
]

pm_tunai_id, pm_bca_id, pm_muamalat_id = str(uuid.uuid4()), str(uuid.uuid4()), str(uuid.uuid4())
payment_methods_data = [
    {"id": pm_tunai_id, "method_name": "Bebas/Tunai", "method_type": "CASH", "bank_id": None},
    {"id": pm_bca_id, "method_name": "EDC/Transfer BCA", "method_type": "BANK", "bank_id": b_bca_id},
    {"id": pm_muamalat_id, "method_name": "EDC-Qris Muamalat", "method_type": "BANK", "bank_id": b_muamalat_id}
]

# Tabel: POS_IMPORT_BATCHES
pos_import_batches = [{
    "id": import_batch_id, "file_name": file_path, "import_date": now, 
    "status": "PROCESSED", "location_id": loc_gbi_id
}]

customers_dict, employees_dict, products_dict = {"ANONIM": str(uuid.uuid4())}, {}, {}
sales_headers, sales_details = [], []
expenses_data = []
stock_movements = []
journal_entries, journal_entry_lines = [], []
stock_balances = {}

def get_product(name):
    if name not in products_dict:
        products_dict[name] = {"id": str(uuid.uuid4()), "int_id": len(products_dict) + 1}
        stock_balances[products_dict[name]["id"]] = 0
    return products_dict[name]

def create_jline(jid, coa_code, entry_type, amount, pay_method_id=None):
    coa = coa_map[coa_code]
    return {
        "id": str(uuid.uuid4()), "journal_id": jid, "coa_id": coa["id"],
        "entry_type": entry_type, "amount": amount, "payment_method_id": pay_method_id
    }

# --- 2. BACA PARETO: CUSTOMERS & EMPLOYEES ---
for sheet in xls.sheet_names:
    if "pareto_karyawan" in sheet.lower() or "pareto_non_member" in sheet.lower():
        df_p = pd.read_excel(file_path, sheet_name=sheet, skiprows=5)
        for _, row in df_p.iterrows():
            if "karyawan" in sheet.lower() and pd.notna(row.iloc[2]):
                emp_name = str(row.iloc[2]).strip()
                if emp_name and emp_name != 'nan' and emp_name not in employees_dict:
                    employees_dict[emp_name] = str(uuid.uuid4())
            elif "non_member" in sheet.lower():
                for col in row.values:
                    val = str(col).strip()
                    if val.startswith("TN ") or val.startswith("NY "):
                        if val not in customers_dict:
                            customers_dict[val] = str(uuid.uuid4())

# --- 3. BACA DATA SALES (PENJUALAN) ---
sales_sheets = [s for s in xls.sheet_names if "penjualan" in s.lower() and "tabular" not in s.lower()]
print(f"\nMemproses Penjualan dari sheets: {sales_sheets}")

for sheet in sales_sheets:
    df_meta = pd.read_excel(file_path, sheet_name=sheet, nrows=10, header=None)
    curr_shift = "1"
    for idx, row in df_meta.iterrows():
        if pd.notna(row.iloc[0]) and "Shift" in str(row.iloc[0]):
            shift_val = str(row.iloc[1]).strip()
            if "1" in shift_val: curr_shift = "1"
            elif "2" in shift_val: curr_shift = "2"
            elif "3" in shift_val: curr_shift = "3"
            break

    df_s = pd.read_excel(file_path, sheet_name=sheet, skiprows=5)
    curr_invoice = None
    curr_sale_id = None
    curr_date = now.date()
    
    for idx, row in df_s.iterrows():
        row_str = " ".join([str(x) for x in row.values if pd.notna(x)])
        
        match_inv = re.search(r'(SL-\d{6}-\d{5})', row_str)
        if match_inv:
            curr_invoice = match_inv.group(1)
            curr_sale_id = str(uuid.uuid4())
            
            date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', row_str)
            if date_match:
                try: curr_date = datetime.datetime.strptime(date_match.group(1), "%m/%d/%Y").date()
                except: pass

            match_pay = re.search(re.escape(curr_invoice) + r'.*?\((.*?)\)', row_str)
            pay_str = match_pay.group(1).strip() if match_pay else "Bebas/Tunai"
            
            sales_headers.append({
                "id": curr_sale_id, "location_id": loc_gbi_id, "employee_id": None, "customer_id": customers_dict["ANONIM"],
                "transaction_date": curr_date, "invoice_number": curr_invoice, "shift": curr_shift, 
                "payment_method": pay_str, "total_hpp": 0.0, "total_amount": 0.0, "total_amount_rounded": 0.0,
                "creation_date": now, "update_date": now, "organization_id": org_id, "last_user": "System"
            })
            
        elif "Nama :" in row_str and curr_sale_id:
            try:
                name_part = row_str.split("Nama :")[1].split("Umur :")[0].strip()
                if name_part in customers_dict:
                    sales_headers[-1]["customer_id"] = customers_dict[name_part]
            except: pass
            
        # PERBAIKAN LOGIKA EXTRAKSI PRODUK & SALES DETAIL
        elif curr_sale_id and pd.notna(row.iloc[0]):
            prod_name = str(row.iloc[0]).strip()
            
            # Pastikan ini baris produk, bukan baris "Diskon Faktur", "Total", atau header "Nama Barang"
            if prod_name and prod_name.upper() not in ["NO", "NAMA BARANG"] and not prod_name.startswith("Diskon") and not prod_name.startswith("Total"):
                try:
                    # Memaksa mengubah teks/angka dari Excel menjadi Float
                    qty = float(str(row.iloc[1]).strip())
                    
                    if qty > 0:
                        p_info = get_product(prod_name)
                        
                        hpp = float(str(row.iloc[2]).strip()) if pd.notna(row.iloc[2]) else 0.0
                        price = float(str(row.iloc[3]).strip()) if pd.notna(row.iloc[3]) else 0.0
                        disc = float(str(row.iloc[4]).strip()) if pd.notna(row.iloc[4]) else 0.0
                        subtot = (qty * price) - disc
                        total_hpp = qty * hpp
                        
                        sales_headers[-1]["total_amount"] += subtot
                        sales_headers[-1]["total_amount_rounded"] += round(subtot, -2)
                        sales_headers[-1]["total_hpp"] += total_hpp
                        
                        sales_details.append({
                            "id": str(uuid.uuid4()), "sale_id": curr_sale_id, "product_id": p_info["id"],
                            "quantity": qty, "quantity_unit": 1, "hpp": hpp, "selling_price": price, "discount": disc,
                            "subtotal": subtot, "creation_date": now, "update_date": now, 
                            "organization_id": org_id, "last_user": "System"
                        })
                        
                        stock_movements.append({
                            "stock_movement_line_number": str(uuid.uuid4()), "reference_type": "SALES_HEADER",
                            "stock_movement_stock_id": p_info["int_id"], "stock_movement_location_id": 1,
                            "stock_movement_purchase_order_number": None, "stock_movement_delivery_note_number": None,
                            "stock_movement_invoice_number": curr_invoice, "stock_movement_date": now,
                            "stock_movement_type": 2, "stock_movement_quantity": qty,
                            "organization_id": org_id, "stock_movement_last_user": "System"
                        })
                        stock_balances[p_info["id"]] -= qty
                except ValueError:
                    # Jika data di kolom Qty tidak bisa diubah jadi angka (misal ada teks aneh), lewati baris ini
                    pass

# TRANSLASI SALES KE JOURNAL
for sale in sales_headers:
    jid = str(uuid.uuid4())
    journal_entries.append({
        "id": jid, "journal_date": sale["transaction_date"], "journal_type": "SALES_HEADER",
        "source_document_id": sale["invoice_number"], "location_id": loc_gbi_id, "import_batch_id": import_batch_id
    })
    
    pay_method = sale["payment_method"].upper()
    pm_id, coa_dr = pm_tunai_id, "1001"
    
    if "BCA" in pay_method: 
        coa_dr, pm_id = "1002", pm_bca_id
    elif "MUAMALAT" in pay_method:
        coa_dr, pm_id = "1002", pm_muamalat_id
        
    journal_entry_lines.append(create_jline(jid, coa_dr, "DR", sale["total_amount"], pm_id))
    journal_entry_lines.append(create_jline(jid, "4001", "CR", sale["total_amount"]))
    journal_entry_lines.append(create_jline(jid, "5001", "DR", sale["total_hpp"]))
    journal_entry_lines.append(create_jline(jid, "1003", "CR", sale["total_hpp"]))

# --- 4. BACA DATA PENGELUARAN KAS (EXPENSES) ---
exp_sheets = [s for s in xls.sheet_names if "pengeluaran" in s.lower()]
for sheet in exp_sheets:
    df_e = pd.read_excel(file_path, sheet_name=sheet, skiprows=5)
    for idx, row in df_e.iterrows():
        if len(row) > 1 and pd.notna(row.iloc[1]) and "KO-" in str(row.iloc[1]):
            try:
                doc_num = str(row.iloc[1]).strip()
                desc = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else "Biaya Operasional"
                raw_amt = str(row.iloc[3])
                
                clean_amt = re.sub(r'[^\d]', '', raw_amt)
                amount = float(clean_amt) if clean_amt else 0.0
                officer = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else "Admin"
                
                if amount > 0:
                    expenses_data.append({
                        "id": str(uuid.uuid4()), "location_id": loc_gbi_id, "expense_date": now.date(),
                        "category": "Operasional", "voucher_number": doc_num, "description": desc,
                        "amount": amount, "officer_name": officer, "creation_date": now, "update_date": now,
                        "organization_id": org_id, "last_user": "System"
                    })
                    
                    jid = str(uuid.uuid4())
                    journal_entries.append({
                        "id": jid, "journal_date": now.date(), "journal_type": "EXPENSE",
                        "source_document_id": doc_num, "location_id": loc_gbi_id, "import_batch_id": import_batch_id
                    })
                    journal_entry_lines.append(create_jline(jid, "6001", "DR", amount, pm_tunai_id))
                    journal_entry_lines.append(create_jline(jid, "1001", "CR", amount, pm_tunai_id))
            except: pass

# --- 5. SIAPKAN DATAFRAME FINAL ---
df_loc = pd.DataFrame(locations_data)
df_coa = pd.DataFrame(coa_data)
df_bank = pd.DataFrame(bank_data)
df_pm = pd.DataFrame(payment_methods_data)
df_batch = pd.DataFrame(pos_import_batches)

df_cust = pd.DataFrame([{"id": cid, "customer_id": i+1, "customer_name": name, "customer_code": f"CUST-{i+1:04d}", "customer_email": None} for i, (name, cid) in enumerate(customers_dict.items())])
df_emp = pd.DataFrame([{"id": eid, "location_id": loc_gbi_id, "employee_id_card": f"EMP-{i+1:03d}", "name": name, "creation_date": now} for i, (name, eid) in enumerate(employees_dict.items())])
df_prod = pd.DataFrame([{"id": info["id"], "product_id": info["int_id"], "product_name": name, "product_code": f"PRD-{info['int_id']:04d}", "product_type_id": 1} for name, info in products_dict.items()])

df_stocks = pd.DataFrame([{
    "id": str(uuid.uuid4()), "stock_id": info["int_id"], "product_id": info["id"], 
    "location_id": loc_gbi_id, "quantity_on_hand": stock_balances[info["id"]], "safety_stock": 10,
    "creation_date": now, "update_date": now, "organization_id": org_id, "last_user": "System"
} for name, info in products_dict.items()])

df_sales_header = pd.DataFrame(sales_headers)

# Memastikan df tidak kosong sebelum diproses
if not df_sales_header.empty:
    excel_formulas = []
    # Looping untuk menulis rumus di setiap baris
    for i in range(len(df_sales_header)):
        # Excel mulai dari baris 2 (baris 1 dipakai untuk judul Header)
        row_excel = i + 2 
        # total_amount ada di Kolom J (kolom ke-10 di df_sales_header)
        formula = f"=ROUND(J{row_excel}, -2)"
        excel_formulas.append(formula)
        
    df_sales_header['rounded_total_amount'] = excel_formulas

# --- 5. REKONSILIASI: MAPPING VS TABULAR REPORT ---
print("\nMelakukan Rekonsiliasi Data Tabular vs Mapping...")
tabular_sheets = [s for s in xls.sheet_names if "tabular" in s.lower()]
df_compare = pd.DataFrame() 

if tabular_sheets:
    sheet_tabular = tabular_sheets[0]
    
    # Cari baris judul tabel (Header) secara dinamis
    df_temp = pd.read_excel(file_path, sheet_name=sheet_tabular, nrows=15, header=None)
    head_idx = 0
    for i, r in df_temp.iterrows():
        if any("Nomor Transaksi" in str(cell) for cell in r.values):
            head_idx = i
            break
            
    df_tab = pd.read_excel(file_path, sheet_name=sheet_tabular, header=head_idx)
    
    # 1. Kumpulkan Data dari Excel Tabular (Fokus: Invoice & Nama Obat)
    tab_records = []
    for _, row in df_tab.iterrows():
        inv = str(row.get('Nomor Transaksi', '')).strip()
        prod = str(row.get('Nama Obat', '')).strip().upper()
        qty = pd.to_numeric(row.get('Qty', 0), errors='coerce')
        
        if pd.notna(qty) and inv and inv != 'nan' and inv != 'None':
            tab_records.append({'Invoice': inv, 'Product': prod, 'Qty_Tabular': qty})
            
    df_tab_clean = pd.DataFrame(tab_records)
    if not df_tab_clean.empty:
        # Jika ada produk yang dientri 2x di 1 struk, gabungkan QTY-nya
        df_tab_clean = df_tab_clean.groupby(['Invoice', 'Product'], as_index=False).sum()
    
    # 2. Kumpulkan Data dari Hasil Mapping Kita (Fokus: Sale ID & Product ID dikembalikan ke string)
    map_records = []
    id_to_prod = {info['id']: name.upper().strip() for name, info in products_dict.items()}
    saleid_to_inv = {s['id']: s['invoice_number'] for s in sales_headers}
    
    for sd in sales_details:
        # Terjemahkan sale_id kembali menjadi Invoice, dan product_id menjadi Nama Obat
        inv = saleid_to_inv.get(sd['sale_id'], '')
        prod = id_to_prod.get(sd['product_id'], '')
        qty = sd['quantity']
        
        map_records.append({'Invoice': inv, 'Product': prod, 'Qty_Mapping': qty})
        
    df_map_clean = pd.DataFrame(map_records)
    if not df_map_clean.empty:
        df_map_clean = df_map_clean.groupby(['Invoice', 'Product'], as_index=False).sum()
    
    # 3. Lakukan FULL OUTER JOIN berdasarkan INVOICE (Sale ID) dan PRODUCT (Nama Obat)
    if not df_tab_clean.empty and not df_map_clean.empty:
        df_compare = pd.merge(df_tab_clean, df_map_clean, on=['Invoice', 'Product'], how='outer')
        
        # Bersihkan nilai kosong (NaN) menjadi 0
        df_compare['Qty_Tabular'] = df_compare['Qty_Tabular'].fillna(0)
        df_compare['Qty_Mapping'] = df_compare['Qty_Mapping'].fillna(0)
        df_compare['Selisih'] = df_compare['Qty_Tabular'] - df_compare['Qty_Mapping']
        
        # 4. Beri Status 
        def check_status(row):
            if row['Qty_Tabular'] > 0 and row['Qty_Mapping'] > 0:
                if row['Selisih'] == 0: return 'MATCH'
                else: return 'BEDA QTY'
            elif row['Qty_Tabular'] > 0 and row['Qty_Mapping'] == 0:
                return 'ADA DI TABULAR, HILANG DI MAPPING'
            else:
                return 'ADA DI MAPPING, HILANG DI TABULAR'
                
        df_compare['Status_Perbandingan'] = df_compare.apply(check_status, axis=1)

        # Rapikan urutan kolom sebelum dicetak
        df_compare = df_compare[['Invoice', 'Product', 'Qty_Tabular', 'Qty_Mapping', 'Selisih', 'Status_Perbandingan']]

datasets = {
    "LOCATIONS": df_loc,
    "CHART_OF_ACCOUNTS": df_coa,
    "BANK": df_bank,
    "PAYMENT_METHODS": df_pm,
    "POS_IMPORT_BATCHES": df_batch,
    "CUSTOMERS": df_cust,
    "EMPLOYEES": df_emp,
    "PRODUCTS": df_prod,
    "SALES_HEADER": df_sales_header,
    "SALES_DETAIL": pd.DataFrame(sales_details),
    "EXPENSES": pd.DataFrame(expenses_data),
    "STOCKS": df_stocks,
    "STOCK_MOVEMENT": pd.DataFrame(stock_movements),
    "JOURNAL_ENTRIES": pd.DataFrame(journal_entries),
    "JOURNAL_ENTRY_LINES": pd.DataFrame(journal_entry_lines),"PERBANDINGAN_TABULAR": df_compare # <--- SHEET BARU DITAMBAHKAN DI SINI
}

# --- 6. SIMPAN KE EXCEL SESUAI ERD ---
output_path = "ERP_Final_Mapped_DBDiagram.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for sheet_name, df in datasets.items():
    ws = wb.create_sheet(title=sheet_name)
    if df.empty: continue
    
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style
        
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

wb.save(output_path)
print(f"\nSelesai! Data berhasil diproses sesuai ERD Final dan disimpan ke: {output_path}")